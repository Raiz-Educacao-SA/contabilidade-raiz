from __future__ import annotations

import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"G:\Drives compartilhados\Financeiro 2026\4. Contábil\2. Rotina\2026\02. DOC_SUPORTE\12. IMOBILIZADO\07 Julho\01.Raiz_Controle-Imobilizado_07.2026.xlsx")
OUTPUT = Path("data/fixed-assets/coligada-01-2026-07-preview.json")


def value(cell):
    item = cell.value
    if isinstance(item, (datetime, date)):
        return item.date().isoformat() if isinstance(item, datetime) else item.isoformat()
    if isinstance(item, Decimal):
        return float(item)
    return item


def normalized_header(item):
    return " ".join(str(item or "").replace("\n", " ").split()).strip()


wb_values = load_workbook(SOURCE, data_only=True, read_only=False, keep_links=True)
wb_formulas = load_workbook(SOURCE, data_only=False, read_only=False, keep_links=True)
assets_ws = wb_values["Cadastro de bens"]
formula_ws = wb_formulas["Cadastro de bens"]
headers = [normalized_header(assets_ws.cell(7, col).value) for col in range(1, assets_ws.max_column + 1)]

assets = []
issues = []
account_totals = defaultdict(lambda: {"items": 0, "cost": 0.0, "accumulated_depreciation": 0.0, "book_value": 0.0})
branch_totals = Counter()

for row in range(8, assets_ws.max_row + 1):
    raw = {headers[col - 1]: value(assets_ws.cell(row, col)) for col in range(1, assets_ws.max_column + 1)}
    if raw.get("Descrição") in (None, "") and raw.get("VALOR CUSTO") in (None, ""):
        continue
    formula_raw = {headers[col - 1]: value(formula_ws.cell(row, col)) for col in range(1, formula_ws.max_column + 1)}
    account = str(raw.get("GRUPO CONTA") or "").strip()
    branch = str(raw.get("FILIAL") or "").strip()
    record = {
        "source_row": row,
        "technical_code": f"CI-01-{row:04d}",
        "company_code": str(raw.get("CÓD.") or "").strip(),
        "branch_code": branch,
        "description": raw.get("Descrição"),
        "invoice_number": raw.get("Número Nota Fiscal"),
        "unit": raw.get("Nome da Unidade"),
        "account_code": account,
        "account_description": raw.get("GRUPO NATUREZA"),
        "note_group": raw.get("NOTA EXPLICATIVA"),
        "cost_center": raw.get("CENTRO CUSTO"),
        "supplier_code": raw.get("CÓDIGO FORNECEDOR"),
        "supplier_name": raw.get("DESCRIÇÃO FORNECEDOR"),
        "acquisition_date": raw.get("DATA AQUISIÇÃO"),
        "disposal_date": raw.get("DATA BAIXA"),
        "quantity": raw.get("QUANTIDADE"),
        "unit_value": raw.get("VALOR UNITARIO"),
        "cost_value": raw.get("VALOR CUSTO"),
        "residual_value": raw.get("VALOR RESIDUAL DESCARTE"),
        "depreciable_value": raw.get("VALOR DEPRECIÁVEL"),
        "accounting_life_months": raw.get("NR. MESES VIDA ÚTIL contábil"),
        "accounting_monthly_depreciation": raw.get("QUOTA DEPREC. MENSAL contábil"),
        "accounting_accumulated_depreciation": raw.get("DEPRECIAÇÃO ACUMULADA contábil"),
        "accounting_book_value": raw.get("SALDO CONTÁBIL"),
        "fiscal_life_months": raw.get("NR. MESES VIDA ÚTIL (fiscal)"),
        "fiscal_monthly_depreciation": raw.get("QUOTA DEPREC. MENSAL fiscal"),
        "fiscal_accumulated_depreciation": raw.get("DEPRECIAÇÃO ACUMULADA fiscal"),
        "fiscal_book_value": raw.get("SALDO FISCAL"),
        "additions": raw.get("Adições"),
        "exclusions": raw.get("Exclusões"),
        "source_account_formula": formula_raw.get("GRUPO CONTA"),
    }
    is_construction_in_progress = account == "1.2.3.01.01"
    record["asset_class"] = "IMOBILIZADO_EM_ANDAMENTO" if is_construction_in_progress else "BEM_DEPRECIAVEL"
    record["opening_book_value"] = record["cost_value"] if is_construction_in_progress else record["accounting_book_value"]
    row_issues = []
    if record["company_code"] not in {"1", "01"}: row_issues.append("COLIGADA_DIVERGENTE")
    if not branch: row_issues.append("FILIAL_AUSENTE")
    if not record["description"]: row_issues.append("DESCRICAO_AUSENTE")
    if not account: row_issues.append("CONTA_AUSENTE")
    if not record["acquisition_date"]: row_issues.append("DATA_AQUISICAO_AUSENTE")
    if not is_construction_in_progress and (not isinstance(record["accounting_life_months"], (int, float)) or record["accounting_life_months"] <= 0): row_issues.append("VIDA_UTIL_CONTABIL_INVALIDA")
    for issue in row_issues:
        issues.append({"source_row": row, "technical_code": record["technical_code"], "issue": issue})
    record["validation_status"] = "REVIEW" if row_issues else "READY"
    assets.append(record)
    branch_totals[branch] += 1
    totals = account_totals[account]
    totals["items"] += 1
    for key, source_key in [("cost", "cost_value"), ("accumulated_depreciation", "accounting_accumulated_depreciation"), ("book_value", "opening_book_value")]:
        number = record[source_key]
        if isinstance(number, (int, float)): totals[key] += float(number)

filiais_ws = wb_values["TabFiliais"]
filiais = []
for row in range(4, filiais_ws.max_row + 1):
    code = value(filiais_ws.cell(row, 1))
    if code not in (None, ""):
        filiais.append({"code": str(code), "name": value(filiais_ws.cell(row, 2)), "cnpj": value(filiais_ws.cell(row, 3))})

tab_contas = wb_values["TabContas"]
account_headers = [normalized_header(tab_contas.cell(3, col).value) for col in range(1, tab_contas.max_column + 1)]
known_account_codes = set()
for row in range(4, tab_contas.max_row + 1):
    for col, header in enumerate(account_headers, 1):
        if "conta" in header.lower() or "cód" in header.lower() or "cod" in header.lower():
            candidate = value(tab_contas.cell(row, col))
            if candidate not in (None, ""):
                known_account_codes.add(str(candidate).strip())

for account in account_totals:
    if account and account not in known_account_codes:
        issues.append({"source_row": None, "technical_code": None, "issue": "CONTA_NAO_LOCALIZADA_TABCONTAS", "account_code": account})

summary_ws = wb_values["Resumo individual"]
summary_reconciliation = []
for row in range(8, summary_ws.max_row + 1):
    account = value(summary_ws.cell(row, 1))
    if account in (None, ""):
        continue
    summary_reconciliation.append({
        "account_code": str(account),
        "account_description": value(summary_ws.cell(row, 2)),
        "items": value(summary_ws.cell(row, 10)),
        "cost": value(summary_ws.cell(row, 11)),
        "accumulated_depreciation": value(summary_ws.cell(row, 15)),
        "control_book_value": value(summary_ws.cell(row, 16)),
        "trial_balance_value": value(summary_ws.cell(row, 18)),
        "difference": value(summary_ws.cell(row, 19)),
        "status": value(summary_ws.cell(row, 17)),
    })

payload = {
    "source": {"file": SOURCE.name, "sha256": hashlib.sha256(SOURCE.read_bytes()).hexdigest(), "reference_date": "2026-07-31"},
    "scope": {"company_code": "01", "all_tabfiliais": True, "status": "PENDING_HOMOLOGATION"},
    "summary": {
        "assets": len(assets),
        "ready": sum(1 for item in assets if item["validation_status"] == "READY"),
        "review": sum(1 for item in assets if item["validation_status"] == "REVIEW"),
        "issues": len(issues),
        "cost": round(sum(float(item["cost_value"] or 0) for item in assets if isinstance(item["cost_value"], (int, float))), 2),
        "accumulated_depreciation": round(sum(float(item["accounting_accumulated_depreciation"] or 0) for item in assets if isinstance(item["accounting_accumulated_depreciation"], (int, float))), 2),
        "book_value": round(sum(float(item["opening_book_value"] or 0) for item in assets if isinstance(item["opening_book_value"], (int, float))), 2),
        "reconciled_accounts": sum(1 for item in summary_reconciliation if abs(float(item["difference"] or 0)) <= 0.01),
        "reconciliation_accounts": len(summary_reconciliation),
    },
    "branches_from_tabfiliais": filiais,
    "asset_count_by_branch": dict(sorted(branch_totals.items())),
    "account_totals": dict(sorted(account_totals.items())),
    "summary_reconciliation": summary_reconciliation,
    "issues": issues,
    "assets": assets,
}
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({"output": str(OUTPUT), **payload["summary"], "branches": filiais, "issue_types": Counter(item["issue"] for item in issues)}, ensure_ascii=False, default=dict, indent=2))
wb_values.close()
wb_formulas.close()
